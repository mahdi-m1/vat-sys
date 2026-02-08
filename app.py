#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
VAT Tax System v3.0.0
=====================
Complete VAT Management System for Bahrain NBR Compliance

Features:
- User Management & Role-Based Access Control
- Invoice Upload & AI-Powered Data Extraction
- VAT Report Generation (NBR Format)
- System Monitoring & Activity Logging
- Ollama AI Integration
"""

import os
import json
import uuid
import hashlib
import logging
from datetime import datetime, timedelta
from functools import wraps
from decimal import Decimal

# Flask & Extensions
from flask import Flask, request, jsonify, send_file, send_from_directory
from flask_cors import CORS
from flask_jwt_extended import (
    JWTManager, create_access_token, create_refresh_token,
    jwt_required, get_jwt_identity, get_jwt
)

# Database
import psycopg2
from psycopg2.extras import RealDictCursor

# Data Processing
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# OCR & Image Processing
import pytesseract
from PIL import Image
from pdf2image import convert_from_path
import io
import base64

# AI Integration
import requests

# Security
import bcrypt

# System Monitoring
import psutil

# Utilities
from werkzeug.utils import secure_filename
from dotenv import load_dotenv

# =====================================================
# CONFIGURATION
# =====================================================

load_dotenv()

app = Flask(__name__, static_folder='static', static_url_path='')
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'vat-secret-key-2026-change-me')
app.config['JWT_SECRET_KEY'] = os.getenv('JWT_SECRET_KEY', 'jwt-secret-key-2026')
app.config['JWT_ACCESS_TOKEN_EXPIRES'] = timedelta(hours=8)
app.config['JWT_REFRESH_TOKEN_EXPIRES'] = timedelta(days=30)
app.config['MAX_CONTENT_LENGTH'] = int(os.getenv('MAX_FILE_SIZE', 10485760))

# CORS Configuration
CORS(app, resources={r"/api/*": {"origins": "*"}})

# JWT Configuration
jwt = JWTManager(app)

# Logging Configuration
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Database Configuration
DB_CONFIG = {
    'host': os.getenv('DB_HOST', 'postgres'),
    'port': os.getenv('DB_PORT', '5432'),
    'database': os.getenv('DB_NAME', 'vat_tax_db'),
    'user': os.getenv('DB_USER', 'vat_user'),
    'password': os.getenv('DB_PASSWORD', 'VatSecure2026!')
}

# Ollama Configuration
OLLAMA_URL = os.getenv('OLLAMA_URL', 'http://192.168.1.225:11434')
OLLAMA_MODEL = os.getenv('OLLAMA_MODEL', 'llama3.2:latest')
OLLAMA_TIMEOUT = int(os.getenv('OLLAMA_TIMEOUT', 60))

# Storage Configuration
STORAGE_PATH = os.getenv('STORAGE_PATH', '/app/storage')
UPLOAD_FOLDER = os.path.join(STORAGE_PATH, 'uploads')
REPORTS_FOLDER = os.path.join(STORAGE_PATH, 'reports')
ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg'}

# Ensure directories exist
for folder in [STORAGE_PATH, UPLOAD_FOLDER, REPORTS_FOLDER]:
    os.makedirs(folder, exist_ok=True)

# =====================================================
# DATABASE HELPERS
# =====================================================

def get_db_connection():
    """Get database connection"""
    try:
        conn = psycopg2.connect(**DB_CONFIG, cursor_factory=RealDictCursor)
        return conn
    except Exception as e:
        logger.error(f"Database connection error: {e}")
        return None

def execute_query(query, params=None, fetch=True):
    """Execute database query"""
    conn = get_db_connection()
    if not conn:
        return None
    try:
        with conn.cursor() as cur:
            cur.execute(query, params)
            if fetch:
                result = cur.fetchall()
            else:
                result = cur.rowcount
            conn.commit()
            return result
    except Exception as e:
        logger.error(f"Query error: {e}")
        conn.rollback()
        return None
    finally:
        conn.close()

def execute_one(query, params=None):
    """Execute query and fetch one result"""
    conn = get_db_connection()
    if not conn:
        return None
    try:
        with conn.cursor() as cur:
            cur.execute(query, params)
            result = cur.fetchone()
            conn.commit()
            return result
    except Exception as e:
        logger.error(f"Query error: {e}")
        conn.rollback()
        return None
    finally:
        conn.close()

# =====================================================
# AUTHENTICATION HELPERS
# =====================================================

def hash_password(password):
    """Hash password using bcrypt"""
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password, hashed):
    """Verify password against hash"""
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def get_user_permissions(user_id):
    """Get user permissions"""
    query = """
        SELECT p.name FROM permissions p
        JOIN role_permissions rp ON p.id = rp.permission_id
        JOIN users u ON u.role_id = rp.role_id
        WHERE u.id = %s
    """
    result = execute_query(query, (user_id,))
    return [r['name'] for r in result] if result else []

def permission_required(permission):
    """Decorator to check user permission"""
    def decorator(f):
        @wraps(f)
        @jwt_required()
        def decorated_function(*args, **kwargs):
            user_id = get_jwt_identity()
            permissions = get_user_permissions(user_id)
            if permission not in permissions and 'admin' not in permissions:
                return jsonify({'error': 'Permission denied'}), 403
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def log_activity(user_id, action, entity_type=None, entity_id=None, details=None, status='success'):
    """Log user activity"""
    try:
        query = """
            INSERT INTO activity_logs (user_id, action, entity_type, entity_id, details, ip_address, status)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """
        ip = request.remote_addr if request else None
        execute_query(query, (user_id, action, entity_type, entity_id, 
                             json.dumps(details) if details else None, ip, status), fetch=False)
    except Exception as e:
        logger.error(f"Failed to log activity: {e}")

# =====================================================
# OLLAMA AI HELPERS
# =====================================================

def get_ollama_settings():
    """Get Ollama settings from database"""
    settings = {}
    query = "SELECT key, value FROM system_settings WHERE key LIKE 'ollama_%'"
    result = execute_query(query)
    if result:
        for row in result:
            settings[row['key']] = row['value']
    return {
        'url': settings.get('ollama_url', OLLAMA_URL),
        'model': settings.get('ollama_model', OLLAMA_MODEL),
        'timeout': int(settings.get('ollama_timeout', OLLAMA_TIMEOUT))
    }

def test_ollama_connection():
    """Test Ollama connection"""
    settings = get_ollama_settings()
    try:
        response = requests.get(f"{settings['url']}/api/version", timeout=5)
        if response.status_code == 200:
            return {'status': 'connected', 'version': response.json()}
        return {'status': 'error', 'message': f"Status code: {response.status_code}"}
    except Exception as e:
        return {'status': 'disconnected', 'message': str(e)}

def get_ollama_models():
    """Get available Ollama models"""
    settings = get_ollama_settings()
    try:
        response = requests.get(f"{settings['url']}/api/tags", timeout=10)
        if response.status_code == 200:
            return response.json().get('models', [])
        return []
    except Exception as e:
        logger.error(f"Failed to get Ollama models: {e}")
        return []

def extract_invoice_data_with_ollama(text):
    """Extract invoice data using Ollama AI"""
    settings = get_ollama_settings()
    
    prompt = f"""You are an expert invoice data extractor. Extract the following information from this invoice text and return ONLY a valid JSON object.

Invoice Text:
{text}

Extract and return this JSON structure (use null for missing values):
{{
    "invoice_number": "string or null",
    "invoice_date": "YYYY-MM-DD or null",
    "vendor_name": "string or null",
    "vendor_vat": "string or null",
    "buyer_name": "string or null", 
    "buyer_vat": "string or null",
    "description": "string or null",
    "amount_exclusive": number or null,
    "vat_amount": number or null,
    "amount_inclusive": number or null,
    "vat_rate": number or null,
    "currency": "string or null",
    "confidence": number between 0 and 100
}}

Return ONLY the JSON object, no explanations."""

    try:
        response = requests.post(
            f"{settings['url']}/api/generate",
            json={
                'model': settings['model'],
                'prompt': prompt,
                'stream': False,
                'options': {'temperature': 0.1}
            },
            timeout=settings['timeout']
        )
        
        if response.status_code == 200:
            result = response.json().get('response', '')
            # Try to parse JSON from response
            try:
                # Find JSON in response
                start = result.find('{')
                end = result.rfind('}') + 1
                if start != -1 and end > start:
                    json_str = result[start:end]
                    return json.loads(json_str)
            except json.JSONDecodeError:
                logger.error(f"Failed to parse Ollama response as JSON")
        return None
    except Exception as e:
        logger.error(f"Ollama extraction error: {e}")
        return None

# =====================================================
# OCR HELPERS
# =====================================================

def extract_text_from_image(image_path):
    """Extract text from image using Tesseract OCR"""
    try:
        image = Image.open(image_path)
        # Configure Tesseract for better results
        custom_config = r'--oem 3 --psm 6 -l eng+ara'
        text = pytesseract.image_to_string(image, config=custom_config)
        return text.strip()
    except Exception as e:
        logger.error(f"OCR error: {e}")
        return ""

def extract_text_from_pdf(pdf_path):
    """Extract text from PDF using OCR"""
    try:
        # Convert PDF to images
        images = convert_from_path(pdf_path, dpi=300)
        text_parts = []
        for image in images:
            custom_config = r'--oem 3 --psm 6 -l eng+ara'
            text = pytesseract.image_to_string(image, config=custom_config)
            text_parts.append(text)
        return '\n'.join(text_parts).strip()
    except Exception as e:
        logger.error(f"PDF OCR error: {e}")
        return ""

def process_uploaded_file(file_path, mime_type):
    """Process uploaded file and extract text"""
    if mime_type == 'application/pdf' or file_path.lower().endswith('.pdf'):
        return extract_text_from_pdf(file_path)
    elif mime_type.startswith('image/') or file_path.lower().endswith(('.jpg', '.jpeg', '.png')):
        return extract_text_from_image(file_path)
    return ""

# =====================================================
# FILE MANAGEMENT HELPERS
# =====================================================

def allowed_file(filename):
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def get_client_storage_path(client_id, year=None):
    """Get storage path for client"""
    client = execute_one("SELECT name FROM clients WHERE id = %s", (client_id,))
    if not client:
        return None
    
    # Sanitize client name for folder
    client_name = secure_filename(client['name'].replace(' ', '_'))
    
    if year:
        path = os.path.join(STORAGE_PATH, 'clients', client_name, str(year))
    else:
        path = os.path.join(STORAGE_PATH, 'clients', client_name)
    
    os.makedirs(path, exist_ok=True)
    return path

def save_uploaded_file(file, client_id, invoice_type, year=None, month=None):
    """Save uploaded file to organized storage"""
    if not year:
        year = datetime.now().year
    if not month:
        month = datetime.now().month
    
    # Get client storage path
    base_path = get_client_storage_path(client_id, year)
    if not base_path:
        return None, "Client not found"
    
    # Create month folder
    month_folder = os.path.join(base_path, f"{month:02d}_{invoice_type}")
    os.makedirs(month_folder, exist_ok=True)
    
    # Generate unique filename
    original_filename = secure_filename(file.filename)
    ext = original_filename.rsplit('.', 1)[1].lower() if '.' in original_filename else 'pdf'
    unique_filename = f"{uuid.uuid4().hex[:8]}_{original_filename}"
    
    file_path = os.path.join(month_folder, unique_filename)
    file.save(file_path)
    
    return file_path, None

# =====================================================
# REPORT GENERATION HELPERS
# =====================================================

def generate_nbr_excel_report(client_id, year, quarter):
    """Generate NBR-compliant Excel report"""
    
    # Get client info
    client = execute_one("SELECT * FROM clients WHERE id = %s", (client_id,))
    if not client:
        return None, "Client not found"
    
    # Get tax period
    period = execute_one("""
        SELECT * FROM tax_periods 
        WHERE client_id = %s AND year = %s AND quarter = %s
    """, (client_id, year, quarter))
    
    # Calculate date range for quarter
    quarter_months = {
        1: (1, 3), 2: (4, 6), 3: (7, 9), 4: (10, 12)
    }
    start_month, end_month = quarter_months[quarter]
    start_date = f"{year}-{start_month:02d}-01"
    end_date = f"{year}-{end_month:02d}-31"
    
    # Get approved sales invoices
    sales_query = """
        SELECT i.*, tt.code as tax_code, tt.name as tax_name, tt.nbr_field_code
        FROM invoices i
        LEFT JOIN tax_types tt ON i.tax_type_id = tt.id
        WHERE i.client_id = %s 
        AND i.invoice_type = 'sales'
        AND i.status = 'approved'
        AND i.invoice_date BETWEEN %s AND %s
        ORDER BY tt.nbr_field_code, i.invoice_date
    """
    sales = execute_query(sales_query, (client_id, start_date, end_date)) or []
    
    # Get approved purchases invoices
    purchases_query = """
        SELECT i.*, tt.code as tax_code, tt.name as tax_name, tt.nbr_field_code
        FROM invoices i
        LEFT JOIN tax_types tt ON i.tax_type_id = tt.id
        WHERE i.client_id = %s 
        AND i.invoice_type = 'purchases'
        AND i.status = 'approved'
        AND i.invoice_date BETWEEN %s AND %s
        ORDER BY tt.nbr_field_code, i.invoice_date
    """
    purchases = execute_query(purchases_query, (client_id, start_date, end_date)) or []
    
    # Create workbook
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers for both sheets
    headers = [
        'VAT return field number',
        'Invoice number',
        'Invoice date',
        'VAT Account Number',
        'Entity Name',
        'Good/Service description',
        'Total BHD (exclusive of VAT)',
        'VAT amount',
        'Total BHD (inclusive of VAT)'
    ]
    
    # ===== SALES SHEET (VP information) =====
    ws_sales = wb.active
    ws_sales.title = "VP information"
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws_sales.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    # Write sales data
    for row_idx, inv in enumerate(sales, 2):
        ws_sales.cell(row=row_idx, column=1, value=inv.get('nbr_field_code', 'L1C1')).border = border
        ws_sales.cell(row=row_idx, column=2, value=inv.get('invoice_number', '')).border = border
        ws_sales.cell(row=row_idx, column=3, value=str(inv.get('invoice_date', ''))).border = border
        ws_sales.cell(row=row_idx, column=4, value=inv.get('counterparty_vat', '')).border = border
        ws_sales.cell(row=row_idx, column=5, value=inv.get('counterparty_name', '')).border = border
        ws_sales.cell(row=row_idx, column=6, value=inv.get('description', 'Sales')).border = border
        ws_sales.cell(row=row_idx, column=7, value=float(inv.get('amount_exclusive', 0) or 0)).border = border
        ws_sales.cell(row=row_idx, column=8, value=float(inv.get('vat_amount', 0) or 0)).border = border
        ws_sales.cell(row=row_idx, column=9, value=float(inv.get('amount_inclusive', 0) or 0)).border = border
    
    # Adjust column widths
    for col in range(1, 10):
        ws_sales.column_dimensions[chr(64 + col)].width = 18
    
    # ===== PURCHASES SHEET (VAT payer information) =====
    ws_purchases = wb.create_sheet("VAT payer information")
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws_purchases.cell(row=1, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    # Write purchases data
    for row_idx, inv in enumerate(purchases, 2):
        ws_purchases.cell(row=row_idx, column=1, value=inv.get('nbr_field_code', 'L8C1')).border = border
        ws_purchases.cell(row=row_idx, column=2, value=inv.get('invoice_number', '')).border = border
        ws_purchases.cell(row=row_idx, column=3, value=str(inv.get('invoice_date', ''))).border = border
        ws_purchases.cell(row=row_idx, column=4, value=inv.get('counterparty_vat', '')).border = border
        ws_purchases.cell(row=row_idx, column=5, value=inv.get('counterparty_name', '')).border = border
        ws_purchases.cell(row=row_idx, column=6, value=inv.get('description', 'Purchases')).border = border
        ws_purchases.cell(row=row_idx, column=7, value=float(inv.get('amount_exclusive', 0) or 0)).border = border
        ws_purchases.cell(row=row_idx, column=8, value=float(inv.get('vat_amount', 0) or 0)).border = border
        ws_purchases.cell(row=row_idx, column=9, value=float(inv.get('amount_inclusive', 0) or 0)).border = border
    
    # Adjust column widths
    for col in range(1, 10):
        ws_purchases.column_dimensions[chr(64 + col)].width = 18
    
    # Generate filename
    client_name = secure_filename(client['name'].replace(' ', '_'))
    filename = f"{client_name}_Q{quarter}_{year}.xlsx"
    
    # Save to reports folder
    client_reports_path = os.path.join(REPORTS_FOLDER, client_name, str(year))
    os.makedirs(client_reports_path, exist_ok=True)
    
    file_path = os.path.join(client_reports_path, filename)
    wb.save(file_path)
    
    # Calculate totals
    total_sales_exclusive = sum(float(inv.get('amount_exclusive', 0) or 0) for inv in sales)
    total_sales_vat = sum(float(inv.get('vat_amount', 0) or 0) for inv in sales)
    total_sales_inclusive = sum(float(inv.get('amount_inclusive', 0) or 0) for inv in sales)
    
    total_purchases_exclusive = sum(float(inv.get('amount_exclusive', 0) or 0) for inv in purchases)
    total_purchases_vat = sum(float(inv.get('vat_amount', 0) or 0) for inv in purchases)
    total_purchases_inclusive = sum(float(inv.get('amount_inclusive', 0) or 0) for inv in purchases)
    
    net_vat = total_sales_vat - total_purchases_vat
    
    return {
        'file_path': file_path,
        'filename': filename,
        'sales_count': len(sales),
        'purchases_count': len(purchases),
        'total_sales_exclusive': total_sales_exclusive,
        'total_sales_vat': total_sales_vat,
        'total_sales_inclusive': total_sales_inclusive,
        'total_purchases_exclusive': total_purchases_exclusive,
        'total_purchases_vat': total_purchases_vat,
        'total_purchases_inclusive': total_purchases_inclusive,
        'net_vat': net_vat
    }, None

# =====================================================
# API ROUTES - AUTHENTICATION
# =====================================================

@app.route('/api/auth/login', methods=['POST'])
def login():
    """User login"""
    data = request.get_json()
    username = data.get('username')
    password = data.get('password')
    
    if not username or not password:
        return jsonify({'error': 'Username and password required'}), 400
    
    user = execute_one("""
        SELECT u.*, r.name as role_name 
        FROM users u 
        JOIN roles r ON u.role_id = r.id 
        WHERE u.username = %s AND u.is_active = TRUE
    """, (username,))
    
    if not user:
        return jsonify({'error': 'Invalid credentials'}), 401
    
    if user['is_locked']:
        return jsonify({'error': 'Account is locked'}), 403
    
    if not verify_password(password, user['password_hash']):
        # Update failed login attempts
        execute_query("""
            UPDATE users SET failed_login_attempts = failed_login_attempts + 1,
            is_locked = CASE WHEN failed_login_attempts >= 4 THEN TRUE ELSE FALSE END
            WHERE id = %s
        """, (user['id'],), fetch=False)
        return jsonify({'error': 'Invalid credentials'}), 401
    
    # Reset failed attempts and update last login
    execute_query("""
        UPDATE users SET failed_login_attempts = 0, last_login = CURRENT_TIMESTAMP
        WHERE id = %s
    """, (user['id'],), fetch=False)
    
    # Get permissions
    permissions = get_user_permissions(user['id'])
    
    # Create tokens
    access_token = create_access_token(
        identity=user['id'],
        additional_claims={'role': user['role_name'], 'permissions': permissions}
    )
    refresh_token = create_refresh_token(identity=user['id'])
    
    log_activity(user['id'], 'login', 'user', user['id'])
    
    return jsonify({
        'access_token': access_token,
        'refresh_token': refresh_token,
        'user': {
            'id': user['id'],
            'username': user['username'],
            'email': user['email'],
            'full_name': user['full_name'],
            'role': user['role_name'],
            'permissions': permissions
        }
    })

@app.route('/api/auth/refresh', methods=['POST'])
@jwt_required(refresh=True)
def refresh():
    """Refresh access token"""
    user_id = get_jwt_identity()
    user = execute_one("SELECT * FROM users WHERE id = %s", (user_id,))
    if not user:
        return jsonify({'error': 'User not found'}), 404
    
    permissions = get_user_permissions(user_id)
    access_token = create_access_token(
        identity=user_id,
        additional_claims={'permissions': permissions}
    )
    return jsonify({'access_token': access_token})

@app.route('/api/auth/me', methods=['GET'])
@jwt_required()
def get_current_user():
    """Get current user info"""
    user_id = get_jwt_identity()
    user = execute_one("""
        SELECT u.id, u.username, u.email, u.full_name, r.name as role
        FROM users u JOIN roles r ON u.role_id = r.id
        WHERE u.id = %s
    """, (user_id,))
    
    if not user:
        return jsonify({'error': 'User not found'}), 404
    
    permissions = get_user_permissions(user_id)
    user['permissions'] = permissions
    return jsonify(user)

@app.route('/api/auth/logout', methods=['POST'])
@jwt_required()
def logout():
    """User logout"""
    user_id = get_jwt_identity()
    log_activity(user_id, 'logout', 'user', user_id)
    return jsonify({'message': 'Logged out successfully'})

# =====================================================
# API ROUTES - USERS MANAGEMENT
# =====================================================

@app.route('/api/users', methods=['GET'])
@permission_required('manage_users')
def get_users():
    """Get all users"""
    users = execute_query("""
        SELECT u.id, u.username, u.email, u.full_name, u.is_active, u.is_locked,
               u.last_login, u.created_at, r.name as role
        FROM users u JOIN roles r ON u.role_id = r.id
        ORDER BY u.created_at DESC
    """)
    return jsonify(users or [])

@app.route('/api/users', methods=['POST'])
@permission_required('manage_users')
def create_user():
    """Create new user"""
    data = request.get_json()
    required = ['username', 'email', 'password', 'role_id']
    
    for field in required:
        if not data.get(field):
            return jsonify({'error': f'{field} is required'}), 400
    
    # Check if username/email exists
    existing = execute_one(
        "SELECT id FROM users WHERE username = %s OR email = %s",
        (data['username'], data['email'])
    )
    if existing:
        return jsonify({'error': 'Username or email already exists'}), 400
    
    password_hash = hash_password(data['password'])
    current_user = get_jwt_identity()
    
    result = execute_one("""
        INSERT INTO users (username, email, password_hash, full_name, role_id, created_by)
        VALUES (%s, %s, %s, %s, %s, %s)
        RETURNING id
    """, (data['username'], data['email'], password_hash, 
          data.get('full_name'), data['role_id'], current_user))
    
    if result:
        log_activity(current_user, 'create_user', 'user', result['id'], {'username': data['username']})
        return jsonify({'id': result['id'], 'message': 'User created successfully'}), 201
    
    return jsonify({'error': 'Failed to create user'}), 500

@app.route('/api/users/<int:user_id>', methods=['PUT'])
@permission_required('manage_users')
def update_user(user_id):
    """Update user"""
    data = request.get_json()
    current_user = get_jwt_identity()
    
    updates = []
    params = []
    
    if 'email' in data:
        updates.append("email = %s")
        params.append(data['email'])
    if 'full_name' in data:
        updates.append("full_name = %s")
        params.append(data['full_name'])
    if 'role_id' in data:
        updates.append("role_id = %s")
        params.append(data['role_id'])
    if 'is_active' in data:
        updates.append("is_active = %s")
        params.append(data['is_active'])
    if 'is_locked' in data:
        updates.append("is_locked = %s")
        params.append(data['is_locked'])
    if 'password' in data and data['password']:
        updates.append("password_hash = %s")
        params.append(hash_password(data['password']))
    
    if not updates:
        return jsonify({'error': 'No fields to update'}), 400
    
    params.append(user_id)
    query = f"UPDATE users SET {', '.join(updates)} WHERE id = %s"
    
    result = execute_query(query, params, fetch=False)
    if result is not None:
        log_activity(current_user, 'update_user', 'user', user_id, data)
        return jsonify({'message': 'User updated successfully'})
    
    return jsonify({'error': 'Failed to update user'}), 500

@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@permission_required('manage_users')
def delete_user(user_id):
    """Delete user (soft delete)"""
    current_user = get_jwt_identity()
    
    if user_id == current_user:
        return jsonify({'error': 'Cannot delete yourself'}), 400
    
    # Check if admin
    user = execute_one("SELECT username FROM users WHERE id = %s", (user_id,))
    if user and user['username'] == 'admin':
        return jsonify({'error': 'Cannot delete admin user'}), 400
    
    result = execute_query(
        "UPDATE users SET is_active = FALSE WHERE id = %s",
        (user_id,), fetch=False
    )
    
    if result is not None:
        log_activity(current_user, 'delete_user', 'user', user_id)
        return jsonify({'message': 'User deleted successfully'})
    
    return jsonify({'error': 'Failed to delete user'}), 500

@app.route('/api/roles', methods=['GET'])
@jwt_required()
def get_roles():
    """Get all roles"""
    roles = execute_query("SELECT * FROM roles ORDER BY id")
    return jsonify(roles or [])

# =====================================================
# API ROUTES - CLIENTS
# =====================================================

@app.route('/api/clients', methods=['GET'])
@jwt_required()
def get_clients():
    """Get all clients"""
    clients = execute_query("""
        SELECT c.*, 
               COUNT(DISTINCT i.id) as invoice_count,
               COUNT(DISTINCT vr.id) as report_count
        FROM clients c
        LEFT JOIN invoices i ON c.id = i.client_id
        LEFT JOIN vat_reports vr ON c.id = vr.client_id
        WHERE c.is_active = TRUE
        GROUP BY c.id
        ORDER BY c.name
    """)
    return jsonify(clients or [])

@app.route('/api/clients/<int:client_id>', methods=['GET'])
@jwt_required()
def get_client(client_id):
    """Get client details"""
    client = execute_one("SELECT * FROM clients WHERE id = %s", (client_id,))
    if not client:
        return jsonify({'error': 'Client not found'}), 404
    
    # Get aliases
    aliases = execute_query(
        "SELECT alias_name FROM client_aliases WHERE client_id = %s",
        (client_id,)
    )
    client['aliases'] = [a['alias_name'] for a in aliases] if aliases else []
    
    return jsonify(client)

@app.route('/api/clients', methods=['POST'])
@permission_required('add_client')
def create_client():
    """Create new client"""
    data = request.get_json()
    
    if not data.get('name') or not data.get('vat_number'):
        return jsonify({'error': 'Name and VAT number are required'}), 400
    
    # Check if VAT number exists
    existing = execute_one(
        "SELECT id FROM clients WHERE vat_number = %s",
        (data['vat_number'],)
    )
    if existing:
        return jsonify({'error': 'VAT number already exists'}), 400
    
    current_user = get_jwt_identity()
    
    result = execute_one("""
        INSERT INTO clients (name, name_ar, vat_number, cr_number, address, phone, email, contact_person, created_by)
        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        RETURNING id
    """, (data['name'], data.get('name_ar'), data['vat_number'], data.get('cr_number'),
          data.get('address'), data.get('phone'), data.get('email'), 
          data.get('contact_person'), current_user))
    
    if result:
        # Create storage folder for client
        get_client_storage_path(result['id'], datetime.now().year)
        
        log_activity(current_user, 'create_client', 'client', result['id'], {'name': data['name']})
        return jsonify({'id': result['id'], 'message': 'Client created successfully'}), 201
    
    return jsonify({'error': 'Failed to create client'}), 500

@app.route('/api/clients/<int:client_id>', methods=['PUT'])
@permission_required('edit_client')
def update_client(client_id):
    """Update client"""
    data = request.get_json()
    current_user = get_jwt_identity()
    
    updates = []
    params = []
    
    fields = ['name', 'name_ar', 'vat_number', 'cr_number', 'address', 'phone', 'email', 'contact_person']
    for field in fields:
        if field in data:
            updates.append(f"{field} = %s")
            params.append(data[field])
    
    if not updates:
        return jsonify({'error': 'No fields to update'}), 400
    
    params.append(client_id)
    query = f"UPDATE clients SET {', '.join(updates)} WHERE id = %s"
    
    result = execute_query(query, params, fetch=False)
    if result is not None:
        log_activity(current_user, 'update_client', 'client', client_id, data)
        return jsonify({'message': 'Client updated successfully'})
    
    return jsonify({'error': 'Failed to update client'}), 500

@app.route('/api/clients/<int:client_id>', methods=['DELETE'])
@permission_required('delete_client')
def delete_client(client_id):
    """Delete client (soft delete)"""
    current_user = get_jwt_identity()
    
    result = execute_query(
        "UPDATE clients SET is_active = FALSE WHERE id = %s",
        (client_id,), fetch=False
    )
    
    if result is not None:
        log_activity(current_user, 'delete_client', 'client', client_id)
        return jsonify({'message': 'Client deleted successfully'})
    
    return jsonify({'error': 'Failed to delete client'}), 500

# =====================================================
# API ROUTES - INVOICES
# =====================================================

@app.route('/api/invoices', methods=['GET'])
@jwt_required()
def get_invoices():
    """Get invoices with filters"""
    client_id = request.args.get('client_id', type=int)
    status = request.args.get('status')
    invoice_type = request.args.get('type')
    
    query = """
        SELECT i.*, c.name as client_name, tt.name as tax_type_name,
               u1.username as uploaded_by_name, u2.username as approved_by_name
        FROM invoices i
        JOIN clients c ON i.client_id = c.id
        LEFT JOIN tax_types tt ON i.tax_type_id = tt.id
        LEFT JOIN users u1 ON i.uploaded_by = u1.id
        LEFT JOIN users u2 ON i.approved_by = u2.id
        WHERE 1=1
    """
    params = []
    
    if client_id:
        query += " AND i.client_id = %s"
        params.append(client_id)
    if status:
        query += " AND i.status = %s"
        params.append(status)
    if invoice_type:
        query += " AND i.invoice_type = %s"
        params.append(invoice_type)
    
    query += " ORDER BY i.created_at DESC LIMIT 500"
    
    invoices = execute_query(query, params if params else None)
    return jsonify(invoices or [])

@app.route('/api/invoices/<int:invoice_id>', methods=['GET'])
@jwt_required()
def get_invoice(invoice_id):
    """Get invoice details"""
    invoice = execute_one("""
        SELECT i.*, c.name as client_name, tt.name as tax_type_name
        FROM invoices i
        JOIN clients c ON i.client_id = c.id
        LEFT JOIN tax_types tt ON i.tax_type_id = tt.id
        WHERE i.id = %s
    """, (invoice_id,))
    
    if not invoice:
        return jsonify({'error': 'Invoice not found'}), 404
    
    return jsonify(invoice)

@app.route('/api/invoices/upload', methods=['POST'])
@permission_required('upload_invoice')
def upload_invoice():
    """Upload and process invoice"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    if not allowed_file(file.filename):
        return jsonify({'error': 'File type not allowed'}), 400
    
    client_id = request.form.get('client_id', type=int)
    invoice_type = request.form.get('invoice_type', 'sales')
    year = request.form.get('year', type=int) or datetime.now().year
    month = request.form.get('month', type=int) or datetime.now().month
    
    if not client_id:
        return jsonify({'error': 'Client ID is required'}), 400
    
    current_user = get_jwt_identity()
    
    # Save file
    file_path, error = save_uploaded_file(file, client_id, invoice_type, year, month)
    if error:
        return jsonify({'error': error}), 400
    
    # Get file info
    file_size = os.path.getsize(file_path)
    mime_type = file.content_type
    
    # Extract text using OCR
    ocr_text = process_uploaded_file(file_path, mime_type)
    
    # Extract data using Ollama AI
    ai_data = None
    confidence = 0
    if ocr_text:
        ai_data = extract_invoice_data_with_ollama(ocr_text)
        if ai_data:
            confidence = ai_data.get('confidence', 50)
    
    # Determine tax type based on VAT rate
    tax_type_id = 1  # Default to VAT 10%
    if ai_data and ai_data.get('vat_rate') is not None:
        vat_rate = ai_data['vat_rate']
        if vat_rate == 0:
            tax_type_id = 2  # VAT 0%
        elif vat_rate >= 0.5:
            tax_type_id = 4  # Excise
    
    # Insert invoice record
    result = execute_one("""
        INSERT INTO invoices (
            client_id, invoice_type, invoice_number, invoice_date,
            counterparty_name, counterparty_vat, description,
            amount_exclusive, vat_amount, amount_inclusive,
            tax_type_id, original_filename, stored_filename, file_path,
            file_size, mime_type, ocr_text, ai_extracted_data,
            confidence_score, status, uploaded_by
        ) VALUES (
            %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s
        ) RETURNING id, uuid
    """, (
        client_id, invoice_type,
        ai_data.get('invoice_number') if ai_data else None,
        ai_data.get('invoice_date') if ai_data else None,
        ai_data.get('vendor_name') if invoice_type == 'purchases' else ai_data.get('buyer_name') if ai_data else None,
        ai_data.get('vendor_vat') if invoice_type == 'purchases' else ai_data.get('buyer_vat') if ai_data else None,
        ai_data.get('description') if ai_data else None,
        ai_data.get('amount_exclusive') if ai_data else None,
        ai_data.get('vat_amount') if ai_data else None,
        ai_data.get('amount_inclusive') if ai_data else None,
        tax_type_id,
        secure_filename(file.filename),
        os.path.basename(file_path),
        file_path,
        file_size,
        mime_type,
        ocr_text,
        json.dumps(ai_data) if ai_data else None,
        confidence,
        'review' if confidence >= 70 else 'pending',
        current_user
    ))
    
    if result:
        log_activity(current_user, 'upload_invoice', 'invoice', result['id'], 
                    {'filename': file.filename, 'client_id': client_id})
        
        return jsonify({
            'id': result['id'],
            'uuid': str(result['uuid']),
            'message': 'Invoice uploaded successfully',
            'extracted_data': ai_data,
            'confidence': confidence,
            'status': 'review' if confidence >= 70 else 'pending'
        }), 201
    
    return jsonify({'error': 'Failed to save invoice'}), 500

@app.route('/api/invoices/<int:invoice_id>', methods=['PUT'])
@permission_required('review_invoice')
def update_invoice(invoice_id):
    """Update invoice data"""
    data = request.get_json()
    current_user = get_jwt_identity()
    
    updates = []
    params = []
    
    fields = ['invoice_number', 'invoice_date', 'counterparty_name', 'counterparty_vat',
              'description', 'amount_exclusive', 'vat_amount', 'amount_inclusive', 
              'tax_type_id', 'review_notes']
    
    for field in fields:
        if field in data:
            updates.append(f"{field} = %s")
            params.append(data[field])
    
    if not updates:
        return jsonify({'error': 'No fields to update'}), 400
    
    updates.append("reviewed_by = %s")
    params.append(current_user)
    updates.append("reviewed_at = CURRENT_TIMESTAMP")
    
    params.append(invoice_id)
    query = f"UPDATE invoices SET {', '.join(updates)} WHERE id = %s"
    
    result = execute_query(query, params, fetch=False)
    if result is not None:
        log_activity(current_user, 'update_invoice', 'invoice', invoice_id, data)
        return jsonify({'message': 'Invoice updated successfully'})
    
    return jsonify({'error': 'Failed to update invoice'}), 500

@app.route('/api/invoices/<int:invoice_id>/approve', methods=['POST'])
@permission_required('approve_invoice')
def approve_invoice(invoice_id):
    """Approve invoice"""
    current_user = get_jwt_identity()
    
    result = execute_query("""
        UPDATE invoices SET status = 'approved', approved_by = %s, approved_at = CURRENT_TIMESTAMP
        WHERE id = %s AND status IN ('pending', 'review')
    """, (current_user, invoice_id), fetch=False)
    
    if result:
        log_activity(current_user, 'approve_invoice', 'invoice', invoice_id)
        return jsonify({'message': 'Invoice approved successfully'})
    
    return jsonify({'error': 'Failed to approve invoice'}), 500

@app.route('/api/invoices/<int:invoice_id>/reject', methods=['POST'])
@permission_required('approve_invoice')
def reject_invoice(invoice_id):
    """Reject invoice"""
    data = request.get_json()
    current_user = get_jwt_identity()
    
    result = execute_query("""
        UPDATE invoices SET status = 'rejected', review_notes = %s,
        reviewed_by = %s, reviewed_at = CURRENT_TIMESTAMP
        WHERE id = %s
    """, (data.get('reason', ''), current_user, invoice_id), fetch=False)
    
    if result:
        log_activity(current_user, 'reject_invoice', 'invoice', invoice_id, {'reason': data.get('reason')})
        return jsonify({'message': 'Invoice rejected'})
    
    return jsonify({'error': 'Failed to reject invoice'}), 500

# =====================================================
# API ROUTES - TAX TYPES
# =====================================================

@app.route('/api/tax-types', methods=['GET'])
@jwt_required()
def get_tax_types():
    """Get all tax types"""
    tax_types = execute_query("SELECT * FROM tax_types WHERE is_active = TRUE ORDER BY id")
    return jsonify(tax_types or [])

# =====================================================
# API ROUTES - REPORTS
# =====================================================

@app.route('/api/reports', methods=['GET'])
@jwt_required()
def get_reports():
    """Get all reports"""
    client_id = request.args.get('client_id', type=int)
    status = request.args.get('status')
    
    query = """
        SELECT vr.*, c.name as client_name,
               u1.username as created_by_name,
               u2.username as approved_by_name
        FROM vat_reports vr
        JOIN clients c ON vr.client_id = c.id
        LEFT JOIN users u1 ON vr.created_by = u1.id
        LEFT JOIN users u2 ON vr.approved_by = u2.id
        WHERE 1=1
    """
    params = []
    
    if client_id:
        query += " AND vr.client_id = %s"
        params.append(client_id)
    if status:
        query += " AND vr.status = %s"
        params.append(status)
    
    query += " ORDER BY vr.created_at DESC"
    
    reports = execute_query(query, params if params else None)
    return jsonify(reports or [])

@app.route('/api/reports/generate', methods=['POST'])
@permission_required('generate_report')
def generate_report():
    """Generate VAT report"""
    data = request.get_json()
    
    client_id = data.get('client_id')
    year = data.get('year')
    quarter = data.get('quarter')
    
    if not all([client_id, year, quarter]):
        return jsonify({'error': 'client_id, year, and quarter are required'}), 400
    
    current_user = get_jwt_identity()
    
    # Generate Excel report
    result, error = generate_nbr_excel_report(client_id, year, quarter)
    if error:
        return jsonify({'error': error}), 400
    
    # Get client name
    client = execute_one("SELECT name FROM clients WHERE id = %s", (client_id,))
    
    # Save report record
    report = execute_one("""
        INSERT INTO vat_reports (
            client_id, report_name, year, quarter,
            total_sales_exclusive, total_sales_vat, total_sales_inclusive,
            total_purchases_exclusive, total_purchases_vat, total_purchases_inclusive,
            net_vat, excel_file_path, status, created_by
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        RETURNING id, uuid
    """, (
        client_id, result['filename'], year, quarter,
        result['total_sales_exclusive'], result['total_sales_vat'], result['total_sales_inclusive'],
        result['total_purchases_exclusive'], result['total_purchases_vat'], result['total_purchases_inclusive'],
        result['net_vat'], result['file_path'], 'draft', current_user
    ))
    
    if report:
        log_activity(current_user, 'generate_report', 'report', report['id'], 
                    {'client_id': client_id, 'year': year, 'quarter': quarter})
        
        return jsonify({
            'id': report['id'],
            'uuid': str(report['uuid']),
            'message': 'Report generated successfully',
            'filename': result['filename'],
            'summary': {
                'sales_count': result['sales_count'],
                'purchases_count': result['purchases_count'],
                'total_sales_vat': result['total_sales_vat'],
                'total_purchases_vat': result['total_purchases_vat'],
                'net_vat': result['net_vat']
            }
        }), 201
    
    return jsonify({'error': 'Failed to save report'}), 500

@app.route('/api/reports/<int:report_id>/approve', methods=['POST'])
@permission_required('approve_report')
def approve_report(report_id):
    """Approve report"""
    current_user = get_jwt_identity()
    
    result = execute_query("""
        UPDATE vat_reports SET status = 'approved', approved_by = %s, approved_at = CURRENT_TIMESTAMP
        WHERE id = %s AND status IN ('draft', 'pending_review')
    """, (current_user, report_id), fetch=False)
    
    if result:
        log_activity(current_user, 'approve_report', 'report', report_id)
        return jsonify({'message': 'Report approved successfully'})
    
    return jsonify({'error': 'Failed to approve report'}), 500

@app.route('/api/reports/<int:report_id>/download', methods=['GET'])
@permission_required('export_report')
def download_report(report_id):
    """Download report file"""
    report = execute_one("SELECT * FROM vat_reports WHERE id = %s", (report_id,))
    
    if not report:
        return jsonify({'error': 'Report not found'}), 404
    
    if not report['excel_file_path'] or not os.path.exists(report['excel_file_path']):
        return jsonify({'error': 'Report file not found'}), 404
    
    current_user = get_jwt_identity()
    log_activity(current_user, 'download_report', 'report', report_id)
    
    return send_file(
        report['excel_file_path'],
        as_attachment=True,
        download_name=report['report_name']
    )

# =====================================================
# API ROUTES - SETTINGS
# =====================================================

@app.route('/api/settings', methods=['GET'])
@permission_required('manage_settings')
def get_settings():
    """Get all settings"""
    settings = execute_query("SELECT key, value, description FROM system_settings ORDER BY key")
    return jsonify(settings or [])

@app.route('/api/settings', methods=['PUT'])
@permission_required('manage_settings')
def update_settings():
    """Update settings"""
    data = request.get_json()
    current_user = get_jwt_identity()
    
    for key, value in data.items():
        execute_query("""
            UPDATE system_settings SET value = %s, updated_by = %s, updated_at = CURRENT_TIMESTAMP
            WHERE key = %s
        """, (value, current_user, key), fetch=False)
    
    log_activity(current_user, 'update_settings', 'settings', None, data)
    return jsonify({'message': 'Settings updated successfully'})

@app.route('/api/settings/ollama', methods=['GET'])
@jwt_required()
def get_ollama_settings_api():
    """Get Ollama settings"""
    settings = get_ollama_settings()
    connection = test_ollama_connection()
    models = get_ollama_models() if connection['status'] == 'connected' else []
    
    return jsonify({
        'settings': settings,
        'connection': connection,
        'available_models': models
    })

@app.route('/api/settings/ollama/test', methods=['POST'])
@jwt_required()
def test_ollama():
    """Test Ollama connection"""
    data = request.get_json()
    url = data.get('url', OLLAMA_URL)
    
    try:
        response = requests.get(f"{url}/api/version", timeout=5)
        if response.status_code == 200:
            return jsonify({'status': 'connected', 'version': response.json()})
        return jsonify({'status': 'error', 'message': f"Status: {response.status_code}"}), 400
    except Exception as e:
        return jsonify({'status': 'disconnected', 'message': str(e)}), 400

# =====================================================
# API ROUTES - MONITORING
# =====================================================

@app.route('/api/monitoring/users', methods=['GET'])
@permission_required('view_logs')
def get_user_monitoring():
    """Get user activity monitoring data"""
    # Active users (last 24 hours)
    active_users = execute_query("""
        SELECT u.id, u.username, u.full_name, r.name as role,
               u.last_login, u.last_activity,
               COUNT(al.id) as actions_today
        FROM users u
        JOIN roles r ON u.role_id = r.id
        LEFT JOIN activity_logs al ON u.id = al.user_id 
            AND al.created_at > CURRENT_TIMESTAMP - INTERVAL '24 hours'
        WHERE u.is_active = TRUE
        GROUP BY u.id, u.username, u.full_name, r.name, u.last_login, u.last_activity
        ORDER BY u.last_activity DESC NULLS LAST
    """)
    
    # User metrics (last 7 days)
    metrics = execute_query("""
        SELECT u.username, u.full_name,
               COALESCE(SUM(um.invoices_uploaded), 0) as invoices_uploaded,
               COALESCE(SUM(um.invoices_approved), 0) as invoices_approved,
               COALESCE(SUM(um.reports_generated), 0) as reports_generated
        FROM users u
        LEFT JOIN user_metrics um ON u.id = um.user_id 
            AND um.date > CURRENT_DATE - INTERVAL '7 days'
        WHERE u.is_active = TRUE
        GROUP BY u.id, u.username, u.full_name
        ORDER BY invoices_uploaded DESC
    """)
    
    return jsonify({
        'active_users': active_users or [],
        'user_metrics': metrics or []
    })

@app.route('/api/monitoring/activity', methods=['GET'])
@permission_required('view_logs')
def get_activity_logs():
    """Get activity logs"""
    limit = request.args.get('limit', 100, type=int)
    user_id = request.args.get('user_id', type=int)
    action = request.args.get('action')
    
    query = """
        SELECT al.*, u.username
        FROM activity_logs al
        LEFT JOIN users u ON al.user_id = u.id
        WHERE 1=1
    """
    params = []
    
    if user_id:
        query += " AND al.user_id = %s"
        params.append(user_id)
    if action:
        query += " AND al.action = %s"
        params.append(action)
    
    query += " ORDER BY al.created_at DESC LIMIT %s"
    params.append(limit)
    
    logs = execute_query(query, params)
    return jsonify(logs or [])

@app.route('/api/monitoring/system', methods=['GET'])
@permission_required('view_dashboard')
def get_system_monitoring():
    """Get system health monitoring data"""
    # System resources
    cpu_percent = psutil.cpu_percent(interval=1)
    memory = psutil.virtual_memory()
    disk = psutil.disk_usage('/')
    
    # Service health checks
    services = {
        'database': {'status': 'healthy' if get_db_connection() else 'down'},
        'ollama': test_ollama_connection()
    }
    
    # Database stats
    db_stats = execute_one("""
        SELECT 
            (SELECT COUNT(*) FROM clients WHERE is_active = TRUE) as clients_count,
            (SELECT COUNT(*) FROM invoices) as invoices_count,
            (SELECT COUNT(*) FROM vat_reports) as reports_count,
            (SELECT COUNT(*) FROM users WHERE is_active = TRUE) as users_count
    """)
    
    # Recent errors
    recent_errors = execute_query("""
        SELECT action, details, created_at
        FROM activity_logs
        WHERE status = 'error'
        ORDER BY created_at DESC
        LIMIT 10
    """)
    
    return jsonify({
        'resources': {
            'cpu_percent': cpu_percent,
            'memory_percent': memory.percent,
            'memory_used_gb': round(memory.used / (1024**3), 2),
            'memory_total_gb': round(memory.total / (1024**3), 2),
            'disk_percent': disk.percent,
            'disk_used_gb': round(disk.used / (1024**3), 2),
            'disk_total_gb': round(disk.total / (1024**3), 2)
        },
        'services': services,
        'database_stats': db_stats,
        'recent_errors': recent_errors or [],
        'uptime': datetime.now().isoformat()
    })

# =====================================================
# API ROUTES - DASHBOARD
# =====================================================

@app.route('/api/dashboard', methods=['GET'])
@jwt_required()
def get_dashboard():
    """Get dashboard data"""
    user_id = get_jwt_identity()
    
    # Summary stats
    stats = execute_one("""
        SELECT 
            (SELECT COUNT(*) FROM clients WHERE is_active = TRUE) as total_clients,
            (SELECT COUNT(*) FROM invoices WHERE status = 'pending') as pending_invoices,
            (SELECT COUNT(*) FROM invoices WHERE status = 'review') as review_invoices,
            (SELECT COUNT(*) FROM invoices WHERE status = 'approved') as approved_invoices,
            (SELECT COUNT(*) FROM vat_reports WHERE status = 'draft') as draft_reports,
            (SELECT COUNT(*) FROM vat_reports WHERE status = 'approved') as approved_reports
    """)
    
    # Recent invoices
    recent_invoices = execute_query("""
        SELECT i.id, i.invoice_number, i.invoice_type, i.status, i.created_at,
               c.name as client_name
        FROM invoices i
        JOIN clients c ON i.client_id = c.id
        ORDER BY i.created_at DESC
        LIMIT 10
    """)
    
    # Recent reports
    recent_reports = execute_query("""
        SELECT vr.id, vr.report_name, vr.year, vr.quarter, vr.status, vr.created_at,
               c.name as client_name
        FROM vat_reports vr
        JOIN clients c ON vr.client_id = c.id
        ORDER BY vr.created_at DESC
        LIMIT 5
    """)
    
    return jsonify({
        'stats': stats,
        'recent_invoices': recent_invoices or [],
        'recent_reports': recent_reports or []
    })

# =====================================================
# STATIC FILES
# =====================================================

@app.route('/')
def serve_index():
    """Serve main page"""
    return send_from_directory(app.static_folder, 'index.html')

@app.route('/<path:path>')
def serve_static(path):
    """Serve static files"""
    if os.path.exists(os.path.join(app.static_folder, path)):
        return send_from_directory(app.static_folder, path)
    return send_from_directory(app.static_folder, 'index.html')

# =====================================================
# ERROR HANDLERS
# =====================================================

@app.errorhandler(404)
def not_found(e):
    return jsonify({'error': 'Not found'}), 404

@app.errorhandler(500)
def server_error(e):
    logger.error(f"Server error: {e}")
    return jsonify({'error': 'Internal server error'}), 500

@jwt.expired_token_loader
def expired_token_callback(jwt_header, jwt_payload):
    return jsonify({'error': 'Token has expired'}), 401

@jwt.invalid_token_loader
def invalid_token_callback(error):
    return jsonify({'error': 'Invalid token'}), 401

# =====================================================
# MAIN
# =====================================================


"""
VAT Tax System v3.0.0 - Missing Endpoints
==========================================
Add these endpoints to your app.py file before the `if __name__ == '__main__':` line
"""

# =====================================================
# HEALTH CHECK ENDPOINT
# =====================================================

@app.route('/api/health', methods=['GET'])
def health_check():
    """Health check endpoint"""
    try:
        # Test database connection
        conn = get_db_connection()
        if conn:
            conn.close()
            db_status = 'connected'
        else:
            db_status = 'disconnected'
        
        return jsonify({
            'status': 'healthy',
            'database': db_status,
            'timestamp': datetime.now().isoformat()
        })
    except Exception as e:
        return jsonify({
            'status': 'unhealthy',
            'error': str(e)
        }), 500


# =====================================================
# DASHBOARD STATS ENDPOINT
# =====================================================

@app.route('/api/dashboard/stats', methods=['GET'])
@jwt_required()
def get_dashboard_stats():
    """Get dashboard statistics"""
    try:
        stats = {}
        
        # Total clients
        result = execute_query("SELECT COUNT(*) as count FROM clients")
        stats['total_clients'] = result[0]['count'] if result else 0
        
        # Total invoices
        result = execute_query("SELECT COUNT(*) as count FROM invoices")
        stats['total_invoices'] = result[0]['count'] if result else 0
        
        # Pending reviews
        result = execute_query("SELECT COUNT(*) as count FROM invoices WHERE status = 'pending_review'")
        stats['pending_reviews'] = result[0]['count'] if result else 0
        
        # Total reports
        result = execute_query("SELECT COUNT(*) as count FROM reports")
        stats['total_reports'] = result[0]['count'] if result else 0
        
        # Recent activity
        result = execute_query("""
            SELECT COUNT(*) as count 
            FROM activity_logs 
            WHERE created_at > NOW() - INTERVAL '24 hours'
        """)
        stats['recent_activity'] = result[0]['count'] if result else 0
        
        return jsonify(stats)
    except Exception as e:
        logger.error(f"Error getting dashboard stats: {e}")
        return jsonify({'error': str(e)}), 500


# =====================================================
# TAX TYPES SETTINGS ENDPOINT
# =====================================================

@app.route('/api/settings/tax-types', methods=['GET'])
@jwt_required()
def get_tax_types_settings():
    """Get all tax types for settings"""
    try:
        tax_types = execute_query("SELECT * FROM tax_types ORDER BY id")
        return jsonify(tax_types or [])
    except Exception as e:
        logger.error(f"Error getting tax types: {e}")
        return jsonify({'error': str(e)}), 500


# =====================================================
# MONITORING ENDPOINTS
# =====================================================

@app.route('/api/monitoring/active-users', methods=['GET'])
@jwt_required()
def get_active_users():
    """Get currently active users"""
    try:
        active_users = execute_query("""
            SELECT DISTINCT u.id, u.username, u.full_name, u.email,
                   MAX(al.created_at) as last_activity
            FROM users u
            LEFT JOIN activity_logs al ON u.id = al.user_id
            WHERE al.created_at > NOW() - INTERVAL '30 minutes'
            GROUP BY u.id, u.username, u.full_name, u.email
            ORDER BY last_activity DESC
        """)
        return jsonify(active_users or [])
    except Exception as e:
        logger.error(f"Error getting active users: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/api/monitoring/user-metrics', methods=['GET'])
@jwt_required()
def get_user_metrics():
    """Get user performance metrics"""
    try:
        metrics = execute_query("""
            SELECT u.id, u.username, u.full_name,
                   COUNT(DISTINCT CASE WHEN al.action = 'upload_invoice' THEN al.id END) as uploads,
                   COUNT(DISTINCT CASE WHEN al.action = 'approve_invoice' THEN al.id END) as approvals,
                   COUNT(DISTINCT CASE WHEN al.action = 'generate_report' THEN al.id END) as reports
            FROM users u
            LEFT JOIN activity_logs al ON u.id = al.user_id
            WHERE al.created_at > NOW() - INTERVAL '30 days' OR al.id IS NULL
            GROUP BY u.id, u.username, u.full_name
            ORDER BY uploads DESC
        """)
        return jsonify(metrics or [])
    except Exception as e:
        logger.error(f"Error getting user metrics: {e}")
        return jsonify({'error': str(e)}), 500


# =====================================================
# OLLAMA SETTINGS UPDATE ENDPOINT
# =====================================================

@app.route('/api/settings/ollama', methods=['PUT'])
@jwt_required()
def update_ollama_settings():
    """Update Ollama AI settings"""
    try:
        data = request.get_json()
        
        # Update environment variables (in production, store in database)
        global OLLAMA_URL, OLLAMA_MODEL
        OLLAMA_URL = data.get('url', OLLAMA_URL)
        OLLAMA_MODEL = data.get('model', OLLAMA_MODEL)
        
        # Log activity
        user_id = get_jwt_identity()
        log_activity(user_id, 'update_ollama_settings', f"Updated Ollama settings: {OLLAMA_URL}, {OLLAMA_MODEL}")
        
        return jsonify({
            'message': 'Settings updated successfully',
            'url': OLLAMA_URL,
            'model': OLLAMA_MODEL
        })
    except Exception as e:
        logger.error(f"Error updating Ollama settings: {e}")
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    app.run(
        host=os.getenv('APP_HOST', '0.0.0.0'),
        port=int(os.getenv('APP_PORT', 5000)),
        debug=os.getenv('APP_DEBUG', 'False').lower() == 'true'
    )
